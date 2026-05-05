"""
工具适配器：桥接各工具页面的后端逻辑到 backend.core 子模块。
"""
import re
import sys
from pathlib import Path

_gui_dir = Path(__file__).resolve().parent.parent
if str(_gui_dir) not in sys.path:
    sys.path.insert(0, str(_gui_dir))

# 软件运行目录：打包后是 exe 所在目录，否则是 gui 目录
if getattr(sys, "frozen", False):
    APP_DIR = Path(sys.executable).resolve().parent
else:
    APP_DIR = _gui_dir

# ---------------------------------------------------------------------------
# 推广链统计 Excel 列索引常量（0-based）
# 来源：巨量广告 → 推广链统计 导出表，列顺序固定
#   col 0: 序号
#   col 1: 计划名（含剧名、留存类型，如"安卓-站内-短剧-每留-真千金一心搞钱-..."）
#   col 2: 操作系统（"Android" / "IOS"）
#   col 3: 页面类型（含"激励"标记时为激励广告）
#   col 7-11: 链接/数据字段（keep_cols 输出列）
# ---------------------------------------------------------------------------
_COL_PLAN_NAME = 1   # 计划名（含剧名 + 留存类型）
_COL_OS        = 2   # 操作系统
_COL_PAGE_TYPE = 3   # 页面类型（含"激励"则为激励行）

from backend.utils.interruptible import (
    StopRequested,
    sleep_ms,
    wait_for_visible,
    wait_for_hidden,
    check_stop,
)


def is_separator_line(text):
    """判断是否为分隔符行"""
    value = (text or "").strip()
    if not value:
        return False
    if re.fullmatch(r"[=═\-—＿_\s]{3,}", value):
        return True
    if re.fullmatch(r"[=═\-—＿_\s]*第\s*\d+\s*组.*[=═\-—＿_\s]*", value):
        return True
    return False


def sanitize_link_text(text):
    """清理链接文本"""
    value = (text or "").replace("\r", "").replace("\n", "").strip()
    value = value.strip("`'\" ")
    value = re.split(r"\s*(?:={3,}|═{3,})", value, maxsplit=1)[0].strip()
    value = value.strip("`'\" ")
    return value


def _normalize_title(title):
    """标题标准化：小写、去空格、只保留中文和字母数字"""
    if not title:
        return ""
    t = title.lower().strip()
    t = re.sub(r'\s+', '', t)
    t = re.sub(r'[^一-龥a-z0-9]', '', t)
    return t


def _fuzzy_find(name, mapping):
    """模糊匹配标题到映射表"""
    key = _normalize_title(name)
    if not key:
        return None
    if key in mapping:
        return mapping[key]
    for k in mapping:
        if key in k or k in key:
            return mapping[k]
    return None


def _parse_material_map(text):
    """解析素材ID映射：格式为 剧名 素材id1 素材id2 ..."""
    m = {}
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        match = re.match(r'^(.*?)(\s+\d{10,}(?:\s+\d{10,})*)$', line)
        if match:
            raw_title = match.group(1).strip()
            ids = match.group(2).strip().split()
            key = _normalize_title(raw_title)
            if key and ids:
                m[key] = {"rawTitle": raw_title, "ids": ids}
    return m


def _parse_judan_map(text):
    """解析剧单修正映射：格式与素材ID映射类似，但值为修正后的剧名"""
    m = {}
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        key = _normalize_title(line)
        if key:
            m[key] = line
    return m


def _strip_drama_prefix(name):
    """剥离剧名前缀：去掉形如 '2026-05-05-每留-' / '2026-05-05-七留-' 等日期+留存类型前缀。

    支持格式：
      YYYY-MM-DD-留存类型-剧名
      YYYY-MM-DD-投手-留存类型-剧名  (如 2026-05-05-每留-超级大单飞)
    留存类型关键词：每留、七留、次留、大单飞、超级
    """
    # 匹配：日期 + 若干前缀段（留存类型/投手名等）+ 剧名
    # 策略：从左到右找到最后一个"已知前缀段"，后面的就是剧名
    m = re.match(
        r'^\d{4}-\d{2}-\d{2}'   # 日期
        r'(?:-[^-]*?(?:每留|七留|次留|留存))'  # 至少一个含留存关键词的段
        r'(?:-[^-]*?(?:每留|七留|次留|留存))*'  # 可能有多个
        r'-',                     # 最后一个 - 分隔
        name
    )
    if m:
        return name[m.end():]
    # 更宽松：日期后跟 1~3 个短段（≤4字符或含已知关键词），后面是剧名
    m2 = re.match(
        r'^\d{4}-\d{2}-\d{2}'
        r'((?:-(?:[^-]{1,4}|[^-]*?(?:每留|七留|次留|留存|激励)[^-]*?))){1,3}'
        r'-',
        name
    )
    if m2:
        return name[m2.end():]
    return name


def _parse_clean_format(text):
    """优先解析格式：按空行分块，http开头为链接，非链接非数字为剧名"""
    dramas = []
    blocks = re.split(r'\n{2,}', text)
    current = None
    for block in blocks:
        block = block.strip()
        if not block:
            continue
        for line in block.splitlines():
            line = line.strip()
            if not line:
                continue
            if is_separator_line(line):
                continue
            clean = sanitize_link_text(line)
            if clean.startswith("http"):
                if current is not None:
                    current["links"].append(clean)
            elif not re.fullmatch(r'\d+', line):
                # 新剧名：剥离日期+留存类型前缀
                drama_name = _strip_drama_prefix(line)
                if current is not None and current["links"]:
                    dramas.append(current)
                current = {"name": drama_name, "links": []}
    if current is not None and current["links"]:
        dramas.append(current)
    return dramas


def _parse_raw_format(text):
    """备用解析：按空格分词，含-的token为剧名，http开头为链接"""
    dramas = []
    current = None
    for token in text.split():
        token = token.strip()
        if not token:
            continue
        if token.startswith("http"):
            clean = sanitize_link_text(token)
            if current is not None:
                current["links"].append(clean)
        elif '-' in token:
            if current is not None and current["links"]:
                dramas.append(current)
            name = _strip_drama_prefix(token)
            current = {"name": name, "links": []}
    if current is not None and current["links"]:
        dramas.append(current)
    return dramas


def _parse_drama_blocks(text):
    """解析短剧数据块，优先 clean 格式，失败则用 raw 格式"""
    dramas = _parse_clean_format(text)
    if not dramas:
        dramas = _parse_raw_format(text)
    return dramas


def do_batch_assign(account_ids_text, dramas_text, ids_per_group, dramas_per_group,
                    material_ids_text='', spacing=1):
    """批量分配 — 完整版"""
    # 1. 解析账户ID
    all_ids = [l.strip() for l in account_ids_text.splitlines()
               if re.match(r'^\d+$', l.strip())]
    if not all_ids:
        return {"ok": False, "error": "未找到有效的账户ID"}

    # 2. 解析短剧数据
    all_dramas = _parse_drama_blocks(dramas_text)
    if not all_dramas:
        return {"ok": False, "error": "未找到有效的短剧数据（需包含剧名和链接）"}

    # 3. 解析素材ID映射
    material_map = {}
    if material_ids_text and material_ids_text.strip():
        material_map = _parse_material_map(material_ids_text)
        if not material_map:
            import logging
            logging.getLogger(__name__).warning(
                "do_batch_assign: material_ids_text 不为空但解析结果为空，请检查格式"
            )

    # 4. 剧单修正（从config读取）
    try:
        from backend.config_manager import load_config
        cfg = load_config()
        titles = cfg.get("common", {}).get("drama_titles", [])
        if titles:
            judan_map = _parse_judan_map("\n".join(titles))
            for drama in all_dramas:
                matched = _fuzzy_find(drama["name"], judan_map)
                if matched:
                    drama["name"] = matched
    except Exception as e:
        import logging
        import traceback
        logging.getLogger(__name__).warning(f"剧单修正失败: {e}\n{traceback.format_exc()}")

    # 5. 分组 + 配对
    ids_per = int(ids_per_group)
    dramas_per = int(dramas_per_group)
    id_groups = [all_ids[i:i+ids_per] for i in range(0, len(all_ids), ids_per)]
    drama_groups = [all_dramas[i:i+dramas_per] for i in range(0, len(all_dramas), dramas_per)]
    pair_count = min(len(id_groups), len(drama_groups))

    if pair_count == 0:
        return {"ok": False, "error": "分组后无法配对（ID组或剧组为空）"}

    # 6. 生成输出
    gap = "\n" * (spacing + 1)  # spacing=1 → 2个换行

    all_contents = []
    for i in range(pair_count):
        ids = id_groups[i]
        dramas = drama_groups[i]
        content = "\n".join(ids) + "\n\n"
        drama_parts = []
        for drama in dramas:
            part = drama["name"] + gap + gap.join(drama["links"])
            # 素材ID匹配
            if material_map:
                mi = _fuzzy_find(drama["name"], material_map)
                if mi and mi["ids"]:
                    part += gap + "\n".join(mi["ids"])
            drama_parts.append(part)
        content += "\n\n\n".join(drama_parts)  # 剧之间3空行
        drama_names_str = "、".join(d["name"] for d in dramas)
        header = f"═══ 第 {i+1} 组（{len(ids)} 个ID，{len(dramas)} 部剧：{drama_names_str}）═══"
        all_contents.append(header + "\n" + content)

    result = ("\n" + "=" * 50 + "\n").join(all_contents)

    return {
        "ok": True,
        "result": result,
        "summary": f"✅ 成功分配 {pair_count} 组（共 {len(all_ids)} 个ID，{len(all_dramas)} 部剧）"
    }


def do_promo_chain(drama_names, directions, log_func, stop_event):
    """推广链生成"""
    from backend.core.promo_chain import run_promotion_chain
    run_promotion_chain(drama_names, directions, log_func, stop_event)


def _normalize_drama_name(name: str) -> str:
    """归一化剧名：去空格、去标点，仅保留中文+字母+数字，小写。"""
    import re as _re
    return _re.sub(r"[^一-龥A-Za-z0-9]", "", (name or "")).lower()


def do_promo_split(mode, log_func, bridge, stop_event, drama_filter=None):
    """推广链分割 — 完整实现：读取Excel，按方向分组，输出结果

    drama_filter: 可选剧名列表（原始字符串），传入则只保留命中清单的行（归一化匹配）。
    """
    try:
        from backend.config_manager import load_config as _lc
        _cfg = _lc()
        _dd = _cfg.get("common", {}).get("download_dir", "")
        src_dir = Path(_dd) if _dd else Path.home() / "Downloads"
    except Exception:
        src_dir = Path.home() / "Downloads"
    keep_cols = [1, 7, 8, 9, 10, 11]
    group_order = ["IOS-每留", "IOS-七留", "Android-每留", "Android-七留"]
    gap_rows = 3

    # 构建过滤集（归一化）
    filter_set = None
    if drama_filter:
        filter_set = {_normalize_drama_name(n) for n in drama_filter if n.strip()}
        log_func(f"🔍 剧名过滤已启用，共 {len(filter_set)} 个剧名（子串包含匹配）")

    log_func(f"📂 处理目录：{src_dir}")
    log_func("📂 匹配文件：推广链统计_*.xlsx")

    try:
        import openpyxl
        from openpyxl import Workbook
    except ImportError:
        log_func("❌ 缺少 openpyxl，请安装: pip install openpyxl")
        return

    import glob as glob_mod
    import os
    from datetime import date

    today_str = date.today().strftime("%Y-%m-%d")
    log_func(f"📅 当前日期：{today_str}")
    pattern = str(Path(src_dir) / f"推广链统计_{today_str}*.xlsx")
    files = glob_mod.glob(pattern)
    skip_markers = ("_处理后", "_processed", "_拆分", "_激励拆分", "_样本")
    files = [f for f in files if not any(m in f for m in skip_markers)]

    # 只保留本工具生成的格式：推广链统计_YYYY-MM-DDTHH-MM-SS.xlsx
    # 使用字符串检查替代正则，性能更好（文件名格式固定）
    def _is_iso_fmt(fname):
        n = os.path.basename(fname)
        return (n.startswith("推广链统计_") and n.endswith(".xlsx")
                and len(n) == len("推广链统计_2000-00-00T00-00-00.xlsx"))
    _rejected = [f for f in files if not _is_iso_fmt(f)]
    files = [f for f in files if _is_iso_fmt(f)]
    if _rejected:
        _sample = ", ".join(os.path.basename(f) for f in _rejected[:3])
        log_func(f"  ℹ️ 已忽略非本工具格式: {_sample}")

    if not files:
        log_func(f"⚠️ 未找到今日（{today_str}）的推广链统计文件")
        log_func(f"   （期望文件名形如：推广链统计_{today_str}T**.xlsx）")
        return

    # 多个文件取最新（按文件名倒序，时间戳在文件名中）
    files = sorted(files, reverse=True)
    log_func(f"📄 找到 {len(files)} 个今日文件，取最新：{os.path.basename(files[0])}")
    files = [files[0]]

    def classify_row(row):
        """按行分类到对应方向"""
        page = str(row[_COL_PAGE_TYPE]) if len(row) > _COL_PAGE_TYPE and row[_COL_PAGE_TYPE] else ""
        if "激励" in page:
            return None
        name = str(row[_COL_PLAN_NAME]) if len(row) > _COL_PLAN_NAME and row[_COL_PLAN_NAME] else ""
        os_v = str(row[_COL_OS])        if len(row) > _COL_OS        and row[_COL_OS]        else ""
        if "每留" in name:
            return os_v + "-每留"
        if "七留" in name:
            return os_v + "-七留"
        return None

    def format_rows(rows):
        """将行数据格式化为tab分隔文本"""
        lines = []
        for row in rows:
            cells = []
            for idx in keep_cols:
                value = row[idx] if len(row) > idx else ""
                cells.append("" if value is None else str(value))
            lines.append("\t".join(cells))
        return "\n".join(lines)

    all_texts = {}
    all_counts = {}

    for src in files:
        fname = os.path.basename(src)
        log_func(f"\n📄 处理: {fname}")

        try:
            wb_src = openpyxl.load_workbook(src)
            ws_src = wb_src.active
            groups = {k: [] for k in group_order}
            _filter_hit = 0
            _filter_miss = 0

            for row in ws_src.iter_rows(values_only=True):
                key = classify_row(row)
                if key in groups:
                    # 剧名过滤：_COL_PLAN_NAME 列是完整计划名（如"安卓-站内-短剧-每留-真千金一心搞钱-lzp-..."）
                    # 用子串包含匹配：filter_set 中任意一项是归一化行名的子串即命中
                    if filter_set is not None:
                        row_drama = _normalize_drama_name(
                            str(row[_COL_PLAN_NAME]) if len(row) > _COL_PLAN_NAME and row[_COL_PLAN_NAME] else ""
                        )
                        if not any(f in row_drama for f in filter_set):
                            _filter_miss += 1
                            continue
                        _filter_hit += 1
                    groups[key].append(row)

            if filter_set is not None:
                log_func(f"  🔍 过滤结果：命中 {_filter_hit} 行 / 跳过 {_filter_miss} 行")

            total = sum(len(v) for v in groups.values())
            if total == 0:
                wb_src.close()
                if filter_set is not None:
                    log_func(f"  ⚠️ 过滤后无匹配数据（剧名清单未命中任何行）")
                else:
                    log_func(f"  ⚠️ 未找到单本推广链数据（可能全为激励数据）")
                continue

            # 保存拆分后的文件
            base = os.path.splitext(fname)[0]
            dst = os.path.join(os.path.dirname(src), base + "_拆分.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.title = "推广链统计"
            for i, key in enumerate(group_order):
                ws.append([key])
                for row in groups[key]:
                    ws.append([row[j] if len(row) > j else None for j in keep_cols])
                if i < len(group_order) - 1:
                    for _ in range(gap_rows):
                        ws.append([])
            try:
                wb.save(dst)
            except PermissionError:
                from datetime import datetime
                ts = datetime.now().strftime('%H%M%S')
                dst = os.path.join(os.path.dirname(src), base + f"_拆分_{ts}.xlsx")
                wb.save(dst)

            wb_src.close()

            # 汇总结果
            counts = {k: len(v) for k, v in groups.items()}
            texts = {k: format_rows(v) for k, v in groups.items()}

            # 合并（多文件时累加）
            for k in group_order:
                all_texts[k] = all_texts.get(k, "") + ("\n" if all_texts.get(k) else "") + texts.get(k, "")
                all_counts[k] = all_counts.get(k, 0) + counts.get(k, 0)

            log_func(f"  ✅ -> {os.path.basename(dst)}")
            for k in group_order:
                log_func(f"     {k}: {counts.get(k, 0)} 条")

        except Exception as e:
            log_func(f"  ❌ 处理失败: {e}")
            import traceback
            log_func(f"  {traceback.format_exc()}")
            traceback.print_exc()

    # 通过 bridge 发送结果到前端
    if all_texts:
        bridge.emit("split-result", {
            "texts": all_texts,
            "counts": all_counts,
            "source_file": os.path.basename(files[0]) if files else "",
        })
        log_func(f"\n✅ 全部完成")
    else:
        log_func(f"\n⚠️ 未产生任何结果")


def do_incentive_split(log_func, bridge, stop_event):
    """激励推广链分割 — 完整实现：读取Excel，按方向分组，输出结果"""
    try:
        from backend.config_manager import load_config as _lc
        _cfg = _lc()
        _dd = _cfg.get("common", {}).get("download_dir", "")
        src_dir = Path(_dd) if _dd else Path.home() / "Downloads"
    except Exception:
        src_dir = Path.home() / "Downloads"
    keep_cols = [1, 7, 8, 9, 10, 11]
    group_order = ["激励-每留", "激励-七留"]
    gap_rows = 3

    log_func(f"📂 处理目录：{src_dir}")
    log_func("📂 匹配文件：推广链统计_*.xlsx")

    try:
        import openpyxl
        from openpyxl import Workbook
    except ImportError:
        log_func("❌ 缺少 openpyxl，请安装: pip install openpyxl")
        return

    import glob as glob_mod
    import os
    from datetime import date

    today_str = date.today().strftime("%Y-%m-%d")
    log_func(f"📅 当前日期：{today_str}")
    pattern = str(Path(src_dir) / f"推广链统计_{today_str}*.xlsx")
    files = glob_mod.glob(pattern)
    skip_markers = ("_处理后", "_processed", "_拆分", "_激励拆分")
    files = [f for f in files if not any(m in f for m in skip_markers)]

    # 只保留本工具生成的格式：推广链统计_YYYY-MM-DDTHH-MM-SS.xlsx
    # 使用字符串检查替代正则，性能更好（文件名格式固定）
    def _is_iso_fmt(fname):
        n = os.path.basename(fname)
        return (n.startswith("推广链统计_") and n.endswith(".xlsx")
                and len(n) == len("推广链统计_2000-00-00T00-00-00.xlsx"))
    _rejected = [f for f in files if not _is_iso_fmt(f)]
    files = [f for f in files if _is_iso_fmt(f)]
    if _rejected:
        _sample = ", ".join(os.path.basename(f) for f in _rejected[:3])
        log_func(f"  ℹ️ 已忽略非本工具格式: {_sample}")

    if not files:
        log_func(f"⚠️ 未找到今日（{today_str}）的激励推广链统计文件")
        log_func(f"   （期望文件名形如：推广链统计_{today_str}T**.xlsx）")
        return

    # 多个文件取最新（按文件名倒序，时间戳在文件名中）
    files = sorted(files, reverse=True)
    log_func(f"📄 找到 {len(files)} 个今日文件，取最新：{os.path.basename(files[0])}")
    files = [files[0]]

    def classify_row(row):
        """按行分类到对应方向（只保留含"激励"的行）"""
        page = str(row[_COL_PAGE_TYPE]) if len(row) > _COL_PAGE_TYPE and row[_COL_PAGE_TYPE] else ""
        if "激励" not in page:
            return None
        name = str(row[_COL_PLAN_NAME]) if len(row) > _COL_PLAN_NAME and row[_COL_PLAN_NAME] else ""
        if "每留" in name:
            return "激励-每留"
        if "七留" in name:
            return "激励-七留"
        return None

    def format_rows(rows):
        """将行数据格式化为tab分隔文本"""
        lines = []
        for row in rows:
            cells = []
            for idx in keep_cols:
                value = row[idx] if len(row) > idx else ""
                cells.append("" if value is None else str(value))
            lines.append("\t".join(cells))
        return "\n".join(lines)

    all_texts = {}
    all_counts = {}

    for src in files:
        fname = os.path.basename(src)
        log_func(f"\n📄 处理: {fname}")

        try:
            wb_src = openpyxl.load_workbook(src)
            ws_src = wb_src.active
            groups = {k: [] for k in group_order}

            for row in ws_src.iter_rows(values_only=True):
                key = classify_row(row)
                if key in groups:
                    groups[key].append(row)

            total = sum(len(v) for v in groups.values())
            if total == 0:
                wb_src.close()
                log_func(f"  ⚠️ 未找到激励数据")
                continue

            # 保存拆分后的文件
            base = os.path.splitext(fname)[0]
            dst = os.path.join(os.path.dirname(src), base + "_激励拆分.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.title = "激励推广链统计"
            for i, key in enumerate(group_order):
                ws.append([key])
                for row in groups[key]:
                    ws.append([row[j] if len(row) > j else None for j in keep_cols])
                if i < len(group_order) - 1:
                    for _ in range(gap_rows):
                        ws.append([])
            try:
                wb.save(dst)
            except PermissionError:
                from datetime import datetime
                ts = datetime.now().strftime('%H%M%S')
                dst = os.path.join(os.path.dirname(src), base + f"_激励拆分_{ts}.xlsx")
                wb.save(dst)

            wb_src.close()

            # 汇总结果
            counts = {k: len(v) for k, v in groups.items()}
            texts = {k: format_rows(v) for k, v in groups.items()}

            # 合并（多文件时累加）
            for k in group_order:
                all_texts[k] = all_texts.get(k, "") + ("\n" if all_texts.get(k) else "") + texts.get(k, "")
                all_counts[k] = all_counts.get(k, 0) + counts.get(k, 0)

            log_func(f"  ✅ -> {os.path.basename(dst)}")
            for k in group_order:
                log_func(f"     {k}: {counts.get(k, 0)} 条")

        except Exception as e:
            log_func(f"  ❌ 处理失败: {e}")
            import traceback
            log_func(f"  {traceback.format_exc()}")
            traceback.print_exc()

    # 通过 bridge 发送结果到前端
    if all_texts:
        bridge.emit("split-result", {
            "texts": all_texts,
            "counts": all_counts,
            "source_file": os.path.basename(files[0]) if files else "",
        })
        log_func(f"\n✅ 全部完成")
    else:
        log_func(f"\n⚠️ 未产生任何结果")

def do_material_push(drama_names, account_id, log_func, stop_event, platform="安卓"):
    """素材搜索推送 — 直接实现 Playwright 自动化（复刻自 SearchDramaMaterialPushToolFrame）"""
    from playwright.sync_api import sync_playwright

    # ─── 常量 ───────────────────────────────────────────
    CDP_URL = "http://127.0.0.1:9222"
    DEFAULT_TIMEOUT = 10_000
    WAIT_AFTER_PUSH = 120_000
    WAIT_AFTER_SEARCH = 5_000
    WAIT_BETWEEN_ROUNDS = 2_000
    MATERIAL_PAGE_URL = (
        "https://www.changdupingtai.com/sale/short-play/manage/material"
        "?material_type=2&audit_status=3&page_index=1&page_size=100"
    )
    SEL_DIALOG = ".arco-drawer, .arco-modal"
    SEL_OPTION = ".arco-select-option"
    SEL_SEARCH_INPUT_ANDROID = 'input[placeholder="请输入素材名称"]'
    SEL_SEARCH_INPUT_IOS = "#search_value_material_input"
    SEL_SEARCH_BTN = "button.distribution_search:has-text('搜索')"
    SEL_TABLE_ROW = "tbody tr.arco-table-tr"

    def _search_input_selector(platform):
        if platform == "iOS":
            return SEL_SEARCH_INPUT_IOS
        return SEL_SEARCH_INPUT_ANDROID

    def _is_browser_system_page(page):
        url = (page.url or "").lower()
        return (
            not url
            or url.startswith("chrome://")
            or url.startswith("devtools://")
            or url.startswith("edge://")
            or url.startswith("about:")
        )

    def _page_has_search_input(page, platform, timeout=2500):
        try:
            page.locator(_search_input_selector(platform)).first.wait_for(
                state="visible", timeout=timeout)
            return True
        except Exception:
            return False

    def _close_blocking_dialogs(page):
        try:
            blocking = page.locator(
                ".arco-drawer-mask, .arco-modal-mask, .arco-drawer, .arco-modal")
            if blocking.count() == 0:
                return
        except Exception:
            return
        log_func("  检测到残留弹窗/遮罩，尝试关闭")
        for _ in range(3):
            try:
                page.keyboard.press("Escape")
                page.wait_for_timeout(500)
            except Exception:
                pass
        close_selectors = [
            ".arco-drawer .arco-drawer-close-icon",
            ".arco-modal .arco-modal-close-icon",
            ".arco-drawer button[aria-label='Close']",
            ".arco-modal button[aria-label='Close']",
            ".arco-drawer .arco-icon-close",
            ".arco-modal .arco-icon-close",
        ]
        for selector in close_selectors:
            try:
                loc = page.locator(selector)
                for i in range(min(loc.count(), 3)):
                    try:
                        loc.nth(i).click(timeout=1_500, force=True)
                        page.wait_for_timeout(500)
                    except Exception:
                        continue
            except Exception:
                continue
        try:
            page.locator(".arco-drawer-mask, .arco-modal-mask").wait_for(
                state="hidden", timeout=3_000)
        except Exception:
            try:
                page.evaluate(
                    """
                    () => {
                        document.querySelectorAll(
                            '.arco-drawer-wrapper, .arco-modal-wrapper, '
                            '.arco-drawer-mask, .arco-modal-mask'
                        ).forEach(el => el.remove());
                        document.body.style.overflow = '';
                    }
                    """
                )
                page.wait_for_timeout(500)
                log_func("  已强制清理残留遮罩")
            except Exception:
                pass

    def _ensure_material_page_ready(page, platform):
        if _page_has_search_input(page, platform, timeout=4_000):
            return True
        try:
            log_func(f"  尝试直接进入素材管理页：{MATERIAL_PAGE_URL}")
            page.goto(MATERIAL_PAGE_URL, wait_until="domcontentloaded", timeout=30_000)
            page.wait_for_timeout(2_000)
            if _page_has_search_input(page, platform, timeout=8_000):
                return True
        except Exception as e:
            log_func(f"  直接进入素材页失败：{e}")
        # 菜单导航备用
        for menu_text in ("推广中心", "短剧列表", "素材管理"):
            try:
                page.get_by_text(menu_text, exact=True).first.click(timeout=3_000)
                page.wait_for_timeout(1_500)
                if _page_has_search_input(page, platform, timeout=3_000):
                    return True
            except Exception:
                continue
        return _page_has_search_input(page, platform, timeout=10_000)

    def _find_material_page(context, platform):
        pages = [pg for pg in context.pages if not _is_browser_system_page(pg)]

        def _score(pg):
            url = (pg.url or "").lower()
            if "changdupingtai.com/sale/short-play/manage/material" in url:
                return 0
            if "manage/material" in url:
                return 1
            if "material" in url:
                return 2
            if "changdupingtai.com" in url:
                return 3
            if "oceanengine" in url:
                return 4
            return 5

        pages.sort(key=_score)
        for pg in pages:
            try:
                # bring_to_front 已移除：素材推送无需置前，force=True 点击不依赖焦点
                pg.set_default_timeout(DEFAULT_TIMEOUT)
                if "changdupingtai.com" in (pg.url or "").lower():
                    if _ensure_material_page_ready(pg, platform):
                        log_func(f"✔ 素材管理页已就绪：{pg.url}")
                        return pg
                    continue
                if _page_has_search_input(pg, platform, timeout=6_000):
                    log_func(f"✔ 已找到包含素材搜索框的页面：{pg.url}")
                    return pg
            except Exception as e:
                log_func(f"  检查页面失败：{e}")
                continue
        # 尝试新建页面
        try:
            log_func(f"  未找到可用页面，尝试打开：{MATERIAL_PAGE_URL}")
            pg = context.new_page()
            pg.set_default_timeout(DEFAULT_TIMEOUT)
            pg.goto(MATERIAL_PAGE_URL, wait_until="domcontentloaded", timeout=30_000)
            # bring_to_front 已移除：新页面无需置前
            if _ensure_material_page_ready(pg, platform):
                log_func(f"✔ 已打开并进入素材管理页：{pg.url}")
                return pg
        except Exception as e:
            log_func(f"  打开素材管理页失败：{e}")
        raise RuntimeError("没有找到素材管理页面，请先在浏览器打开素材管理页后再开始推送")

    def _search_drama(page, name, platform):
        try:
            page.keyboard.press("Escape")
        except Exception:
            pass
        search_input = page.locator(_search_input_selector(platform)).first
        wait_for_visible(search_input, DEFAULT_TIMEOUT, stop_event)
        search_input.click()
        search_input.fill("")
        search_input.fill(name)
        log_func(f"  ✔ 已输入剧名：{name}")
        search_btn = page.locator(SEL_SEARCH_BTN).first
        wait_for_visible(search_btn, DEFAULT_TIMEOUT, stop_event)
        button_text = search_btn.inner_text(timeout=DEFAULT_TIMEOUT).strip()
        if "搜索" not in button_text:
            raise RuntimeError(f"搜索按钮定位异常，当前命中按钮文本：{button_text}")
        search_btn.click()
        log_func("  ✔ 已点击搜索按钮")
        sleep_ms(page, WAIT_AFTER_SEARCH, stop_event)

    def _verify_search_result(page, name):
        rows = page.locator(SEL_TABLE_ROW)
        count = rows.count()
        if count == 0:
            log_func(f"  ⚠️ 搜索结果为空，跳过「{name}」")
            return False
        try:
            row_text = rows.nth(0).inner_text().strip()
            snippet = row_text.replace("\n", " ")[:60]
            if name in row_text:
                log_func(f"  ✔ 搜索结果验证通过（首行：{snippet}...）")
                return True
            log_func(f"  ⚠️ 首行不含「{name}」：{snippet}...，跳过")
            return False
        except Exception as e:
            log_func(f"  ⚠️ 验证搜索结果出错：{e}，跳过")
            return False

    def _push_all(page, name, account_id):
        # 切换分页为 100 条/页
        try:
            page_size_sel = page.locator(".arco-pagination-option .arco-select-view")
            if page_size_sel.count() > 0 and "100" not in page_size_sel.first.inner_text(timeout=3000):
                page_size_sel.first.click()
                sleep_ms(page, 500, stop_event)
                opt_100 = page.locator(".arco-select-option:visible", has_text="100 条/页")
                if opt_100.count() > 0:
                    opt_100.first.click()
                    sleep_ms(page, 2000, stop_event)
                    log_func("  ✔ 已切换分页为 100 条/页")
                else:
                    page.keyboard.press("Escape")
            else:
                log_func("  ✔ 分页已是 100 条/页")
        except StopRequested:
            raise
        except Exception as e:
            log_func(f"  ⚠️ 切换分页失败：{e}，继续执行")
        # 全选素材
        checkbox = page.locator("table thead .arco-checkbox-mask")
        wait_for_visible(checkbox, DEFAULT_TIMEOUT, stop_event)
        checkbox.click()
        log_func("  ✔ 已点击全选框")
        batch_btn = page.get_by_role("button", name="批量操作")
        wait_for_visible(batch_btn, DEFAULT_TIMEOUT, stop_event)
        batch_btn.click()
        log_func("  ✔ 已点击批量操作")
        push_btn = page.get_by_text("批量推送", exact=True)
        wait_for_visible(push_btn, DEFAULT_TIMEOUT, stop_event)
        push_btn.click()
        log_func("  ✔ 已点击批量推送")
        dialog = page.locator(SEL_DIALOG).filter(has_text="批量素材推送")
        wait_for_visible(dialog, DEFAULT_TIMEOUT, stop_event)
        log_func("  ✔ 弹窗已出现")
        dialog.locator(".arco-select", has=page.get_by_placeholder("请选择媒体渠道")).click()
        opt_engine = page.locator(SEL_OPTION, has_text="巨量引擎")
        wait_for_visible(opt_engine, DEFAULT_TIMEOUT, stop_event)
        opt_engine.click()
        log_func("  ✔ 已选择媒体渠道：巨量引擎")
        dialog.locator(".arco-select", has=page.get_by_placeholder("请选择投放产品")).click()
        opt_product = page.locator(SEL_OPTION, has_text="红果免费短剧")
        wait_for_visible(opt_product, DEFAULT_TIMEOUT, stop_event)
        opt_product.click()
        log_func("  ✔ 已选择投放产品：红果免费短剧")
        page.keyboard.press("Escape")
        account_input = dialog.locator("#ad_account_ids_input")
        wait_for_visible(account_input, DEFAULT_TIMEOUT, stop_event)
        account_input.fill(account_id)
        log_func(f"  ✔ 已输入广告账户 ID：{account_id}")
        confirm_btn = dialog.get_by_role("button", name="确定")
        wait_for_visible(confirm_btn, DEFAULT_TIMEOUT, stop_event)
        confirm_btn.click()
        log_func("  ✔ 已点击确定，等待推送完成...")
        wait_for_hidden(dialog, WAIT_AFTER_PUSH, stop_event)
        log_func(f"  ✅ 「{name}」推送完成（弹窗已自动关闭）")

    # ─── 主流程 ──────────────────────────────────────────
    log_func(f"🔍 开始推送 {len(drama_names)} 个剧名，账户: {account_id}，方向: {platform}")
    try:
        with sync_playwright() as p:
            browser = p.chromium.connect_over_cdp(CDP_URL)
            context = browser.contexts[0]
            page = _find_material_page(context, platform)
            # bring_to_front 已移除：无需置前
            page.set_default_timeout(DEFAULT_TIMEOUT)
            log_func(f"✔ 使用页面：{page.url}")

            success_count = 0
            skip_count = 0
            fail_count = 0
            failed_names = []

            for idx, name in enumerate(drama_names, 1):
                check_stop(stop_event)
                log_func("─" * 50)
                log_func(f"🔄 [{idx}/{len(drama_names)}] 处理剧名：{name}")
                try:
                    _search_drama(page, name, platform)
                    if not _verify_search_result(page, name):
                        skip_count += 1
                        failed_names.append(name)
                        _close_blocking_dialogs(page)
                        sleep_ms(page, WAIT_BETWEEN_ROUNDS, stop_event)
                        continue
                    _push_all(page, name, account_id)
                    success_count += 1
                    sleep_ms(page, WAIT_BETWEEN_ROUNDS, stop_event)
                except StopRequested:
                    raise
                except Exception as e:
                    log_func(f"  ❌ 处理「{name}」失败：{e}")
                    fail_count += 1
                    failed_names.append(name)
                    try:
                        _close_blocking_dialogs(page)
                        page.keyboard.press("Escape")
                        sleep_ms(page, WAIT_BETWEEN_ROUNDS, stop_event)
                        wait_for_hidden(
                            page.locator(SEL_DIALOG).filter(has_text="批量素材推送"),
                            DEFAULT_TIMEOUT, stop_event)
                    except StopRequested:
                        raise
                    except Exception:
                        pass

            log_func("=" * 50)
            log_func(f"🎉 全部完成！共处理 {len(drama_names)} 个剧名")
            log_func(f"   ✅ 成功推送：{success_count} 个")
            log_func(f"   ⚠️ 跳过（搜索无结果或不匹配）：{skip_count} 个")
            log_func(f"   ❌ 失败：{fail_count} 个")
            log_func("=" * 50)
            if failed_names:
                log_func("📋 失败/跳过的剧名清单：")
                for n in failed_names:
                    log_func(n)

    except StopRequested:
        log_func("⏹ 已停止")
    except Exception as e:
        import traceback
        log_func(f"❌ Playwright 连接失败: {e}")
        log_func(traceback.format_exc())
    finally:
        try:
            browser.close()  # noqa: F821
        except (NameError, Exception):
            pass


def do_incentive_chain(count, suffix, log_func, stop_event):
    """激励推广链生成"""
    from backend.core.incentive_tools import run_incentive_promo_chain
    log_func(f"🚀 开始生成激励推广链: {count} 次, 后缀: {suffix}")
    run_incentive_promo_chain(count, suffix, log_func, stop_event)


def do_incentive_push(account_id, log_func, stop_event):
    """激励素材推送"""
    from backend.core.incentive_tools import run_incentive_push
    run_incentive_push(account_id, log_func, stop_event)


def do_incentive_link_assign(params):
    """激励链接分配"""
    raw_data = params.get("raw_data", "")
    account_ids = params.get("account_ids", "")
    ids_per_group = params.get("ids_per_group", 6)

    if not raw_data.strip() or not account_ids.strip():
        return {"ok": False, "error": "请输入原始数据和账户ID"}

    ids = [line.strip().split('\t')[0] for line in account_ids.strip().splitlines() if line.strip()]
    data_rows = [line.strip() for line in raw_data.strip().splitlines() if line.strip()]

    id_groups = [ids[i:i+ids_per_group] for i in range(0, len(ids), ids_per_group)]

    lines = []
    for g_idx, group_ids in enumerate(id_groups):
        data_row = data_rows[g_idx] if g_idx < len(data_rows) else ""
        # 提取组标签（Tab分隔的第一个字段）作为组名
        parts = data_row.split('\t') if data_row else []
        group_label = parts[0].strip() if parts else f"组{g_idx + 1}"
        lines.append(f"{'═'*20} {group_label} {'═'*20}")
        lines.append('\n'.join(group_ids))
        if data_row:
            lines.append('')
            lines.append(data_row.replace('\t', '\n'))
        lines.append('')

    return {
        "ok": True,
        "result": '\n'.join(lines),
        "summary": f"✅ {len(ids)} 个账户分为 {len(id_groups)} 组"
    }


def do_crawl_material(drama_names, min_cost, min_count, log_func, stop_event):
    """爬取历史跑量素材ID — 调用 D:\\红果\\爬素材id.py 的逻辑"""
    import asyncio
    if sys.platform == "win32":
        loop = asyncio.SelectorEventLoop()
    else:
        loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)

    # 加载爬素材脚本：优先用内置模块（打包后随软件分发），找不到再回退到外部脚本
    crawl_mod = None
    try:
        from backend import crawl_material_script as _crawl_mod
        crawl_mod = _crawl_mod
        log_func("📦 使用内置爬素材脚本")
    except Exception as e:
        log_func(f"⚠️ 内置爬素材脚本加载失败，尝试外部脚本：{e}")
        import importlib.util
        # 外部脚本路径：从配置读取或使用默认相对路径，避免硬编码绝对路径
        from backend.config_manager import load_config
        cfg = load_config()
        script_path_str = cfg.get("common", {}).get("crawl_script_path", "")
        if script_path_str:
            script_path = Path(script_path_str)
        else:
            script_path = Path(__file__).resolve().parent.parent / "爬素材id.py"
        if not script_path.exists():
            log_func(f"❌ 外部脚本也不存在: {script_path}")
            return
        try:
            spec = importlib.util.spec_from_file_location("crawl_material", str(script_path))
            crawl_mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(crawl_mod)
        except Exception as e2:
            log_func(f"❌ 加载外部脚本失败: {e2}")
            return

    from playwright.sync_api import sync_playwright

    log_func(f"🚀 开始爬取 {len(drama_names)} 个剧名，最低消耗: {min_cost}元，最少集数: {min_count}")

    results = []

    try:
        with sync_playwright() as p:
            browser = p.chromium.connect_over_cdp("http://127.0.0.1:9222")
            page = crawl_mod.get_target_page(browser)
            # bring_to_front 已移除：爬素材无需置前

            log_func(f"✅ 已连接页面: {page.url}")

            for index, drama_name in enumerate(drama_names, start=1):
                check_stop(stop_event)

                try:
                    log_func(f"\n{'='*40}")
                    log_func(f"📖 处理第 {index}/{len(drama_names)} 个: {drama_name}")

                    crawl_mod.clear_and_search(page, drama_name, stop_event)

                    all_items = []
                    seen_nums = set()
                    page_index = 1
                    # 页面已按消耗降序，每页约20条，抓够去重后的集数即可停止
                    max_pages = max(3, (min_count // 20) + 2)  # 留点余量

                    while page_index <= max_pages:
                        check_stop(stop_event)
                        log_func(f"  📄 抓取第 {page_index} 页...")
                        page_items = crawl_mod.collect_from_current_page(page, drama_name)
                        all_items.extend(page_items)

                        # 统计去重后的集数，够了就不翻页了
                        for item in page_items:
                            seen_nums.add(item["num"])
                        if len(seen_nums) >= min_count:
                            log_func(f"  ✅ 已收集 {len(seen_nums)} 个不同集数，够了")
                            break

                        if not crawl_mod.go_next_page(page, stop_event):
                            log_func(f"  📄 共抓取 {page_index} 页")
                            break
                        page_index += 1

                    all_numbers = crawl_mod.filter_items(all_items, min_cost=min_cost, min_count=min_count)

                    results.append({
                        "drama_name": drama_name,
                        "numbers": all_numbers
                    })

                    line = f"{drama_name} {' '.join(all_numbers)}".strip()
                    log_func(f"  ✅ 结果: {line}")

                except StopRequested:
                    raise
                except Exception as e:
                    import traceback
                    log_func(f"  ❌ 剧名【{drama_name}】处理失败: {e}")
                    log_func(f"  {traceback.format_exc()}")
                    results.append({"drama_name": drama_name, "numbers": []})
                    continue

            browser.close()

    except StopRequested:
        log_func("⏹ 已停止")
    except Exception as e:
        import traceback as _tb
        log_func(f"❌ Playwright 连接失败: {e}")
        log_func(_tb.format_exc())
        return
    finally:
        try:
            browser.close()  # noqa: F821
        except (NameError, Exception):
            pass

    # 汇总结果
    sorted_results = sorted(results, key=lambda x: (len(x["numbers"]) < min_count, -len(x["numbers"])))
    all_lines = []
    for item in sorted_results:
        line = f'{item["drama_name"]} {" ".join(item["numbers"])}'.strip()
        all_lines.append(line)

    result_text = "\n".join(all_lines)
    log_func(f"\n{'='*40}")
    log_func(f"✅ 全部完成！共 {len(drama_names)} 个剧名")
    log_func(f"RESULT:{result_text}")

    # 同时保存到文件（保存到软件运行目录）
    try:
        output_path = APP_DIR / "批量抓取结果.txt"
        output_path.write_text(result_text, encoding="utf-8")
        log_func(f"📁 结果已保存: {output_path}")
    except Exception as e:
        log_func(f"⚠️ 保存文件失败: {e}")


def parse_and_add_to_profile(profile_key, result_text):
    """解析批量分配结果，写入config.json的指定profile的groups"""
    from backend.config_manager import load_config, save_config

    cfg = load_config()
    profile = cfg.get("profiles", {}).get(profile_key)
    if profile is None:
        return {"ok": False, "error": f"配置中不存在 profile: {profile_key}"}

    # 按 ═{3,} 或 ={3,} 分隔符拆分组
    chunks = re.split(r'\s*\n\s*={3,}\s*\n\s*', result_text.strip())
    groups = []

    for chunk in chunks:
        chunk = chunk.strip()
        if not chunk:
            continue

        # 清理：把 ═══ 类分隔符替换为换行
        chunk = re.sub(r'(?:={3,}|═{3,})', '\n', chunk)

        lines = []
        for line in chunk.splitlines():
            clean = line.strip().strip("`'\" ")
            if not clean or is_separator_line(clean):
                continue
            lines.append(clean)

        if not lines:
            continue

        # 前面的纯数字行 → 账户ID
        account_ids = []
        idx = 0
        while idx < len(lines):
            if lines[idx].isdigit():
                account_ids.append(lines[idx])
                idx += 1
            else:
                break

        if not account_ids:
            continue

        # 后面的行 → 剧名 + 链接 + 素材ID
        drama_list = []
        current = None

        def _is_url(s):
            return bool(re.match(r'^`?\s*https?://', s, re.I))

        def _classify_link(url):
            u = url.lower()
            if "action_type=effective_play" in u: return "video"
            if "action_type=click" in u: return "click"
            if "action_type=view" in u: return "show"
            if "effective_play" in u: return "video"
            if "/display/" in u: return "click"
            if "impression" in u: return "show"
            return "unknown"

        def _is_material_ids_line(s):
            parts = s.split()
            return bool(parts) and len(parts) > 1 and all(p.isdigit() for p in parts)

        def flush():
            nonlocal current
            if current:
                drama_list.append(current)
                current = None

        while idx < len(lines):
            clean = lines[idx].strip().strip("`'\" ")
            if not clean or is_separator_line(clean):
                idx += 1
                continue

            # URL → 链接
            if _is_url(clean):
                if current is not None:
                    link = sanitize_link_text(clean)
                    link_type = _classify_link(link)
                    if link_type != "unknown" and not current.get(link_type):
                        current[link_type] = link
                idx += 1
                continue

            # 素材ID行（多个数字空格分隔）
            if current and _is_material_ids_line(clean):
                current["material_ids"] = clean.split()
                idx += 1
                continue

            # 单个纯数字 → 可能是素材ID的一部分
            if clean.isdigit() and current:
                current.setdefault("material_ids", []).append(clean)
                idx += 1
                continue

            # 其他 → 新剧名
            flush()
            current = {"name": clean, "click": "", "show": "", "video": "", "material_ids": []}
            idx += 1

        flush()

        group = {
            "id": len(groups) + 1,  # 批量导入时按顺序分配稳定 id
            "account_ids": account_ids,
            "dramas": drama_list,
        }
        groups.append(group)

    if not groups:
        return {"ok": False, "error": "未能解析出有效的组数据"}

    profile["groups"] = groups
    save_config(cfg)
    return {"ok": True, "count": len(groups)}


def add_incentive_groups_to_profile(profile_key, result_text):
    """解析激励链接分配结果，写入config.json的指定profile的groups"""
    from backend.config_manager import load_config, save_config

    cfg = load_config()
    profile = cfg.get("profiles", {}).get(profile_key)
    if profile is None:
        return {"ok": False, "error": f"配置中不存在 profile: {profile_key}"}

    # 先提取每个分隔行中的组名
    group_names = re.findall(r'═{5,}\s*(.+?)\s*═{5,}', result_text)
    # 按组分隔符拆分
    blocks = re.split(r'═{5,}[^═]*═{5,}', result_text)
    groups = []
    valid_block_idx = 0
    for block in blocks:
        block = block.strip()
        if not block:
            continue
        lines = [l.strip() for l in block.splitlines() if l.strip()]
        if not lines:
            continue

        account_ids = []
        link_lines = []
        for line in lines:
            if re.match(r'^\d+$', line):
                account_ids.append(line)
            elif re.match(r'^https?://', line, re.I):
                link_lines.append(line)
            # 跳过非数字、非链接的行（如组名标签）

        if account_ids:
            # 根据链接内容自动识别类型，不依赖顺序
            click_url = ""
            show_url = ""
            play_url = ""
            for link in link_lines:
                low = link.lower()
                if not click_url and ("action_type=click" in low or "/display/" in low):
                    click_url = link
                elif not play_url and "action_type=effective_play" in low:
                    play_url = link
                elif not show_url and ("action_type=view" in low or "impression" in low):
                    show_url = link
            # 兜底：如果无法识别类型，按顺序分配剩余未分配的链接
            unassigned = [l for l in link_lines if l not in (click_url, show_url, play_url)]
            if not click_url and unassigned:
                click_url = unassigned.pop(0)
            if not show_url and unassigned:
                show_url = unassigned.pop(0)
            if not play_url and unassigned:
                play_url = unassigned.pop(0)

            # 从分隔行提取"组X"（如 "2026-05-01-组1-每留" → "组1"）
            raw_label = group_names[valid_block_idx] if valid_block_idx < len(group_names) else ""
            m = re.search(r'组\d+', raw_label)
            gname = m.group(0) if m else f"组{valid_block_idx + 1}"

            group = {
                "id": len(groups) + 1,  # 批量导入时按顺序分配稳定 id
                "account_ids": account_ids,
                "group_name": gname,
                "click_url": click_url,
                "show_url": show_url,
                "play_url": play_url,
            }
            groups.append(group)
            valid_block_idx += 1

    if not groups:
        return {"ok": False, "error": "未能解析出有效的组数据"}

    profile["groups"] = groups
    save_config(cfg)
    return {"ok": True, "count": len(groups)}


def do_rta_set(drama_type: str, aadvids: list, log_func) -> None:
    """RTA 设置 — 批量设置生效范围"""
    from backend.tools.rta_set import do_rta_set as _do_rta_set
    _do_rta_set(drama_type, aadvids, log_func)


def do_rta_check(drama_type: str, aadvids: list, log_func) -> None:
    """RTA 检测 — 批量检测启用状态"""
    from backend.tools.rta_check import do_rta_check as _do_rta_check
    _do_rta_check(drama_type, aadvids, log_func)
